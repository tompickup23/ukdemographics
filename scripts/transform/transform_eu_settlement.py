#!/usr/bin/env python3
"""
Aggregate the Home Office EU Settlement Scheme dataset for UKD.

The EUSS is the single largest ongoing immigration system process by
volume but is rarely surfaced because it grants confirmed status to
people already in the UK, so it does NOT count as 'immigration' in
the LTIM sense. It still affects population denominators and is
politically live as 371,000 settled-status grants were issued in
the year ending March 2026 (12 percent higher than YE Mar 2025).

Input:
  data/raw/ho_visas_extra/eu-settlement-scheme-datasets-mar-2026.xlsx

Output:
  src/data/live/eu-settlement-scheme.json

Sheets used:
  Data_EUSS_D02: concluded applications by outcome year + outcome
  Data_EUSS_D03: concluded applications by outcome + application type

Source: Home Office Immigration Statistics, year ending March 2026
release (21 May 2026).
"""
import json
from collections import defaultdict
import sys
from pathlib import Path

import openpyxl

# Release files carry their period in the filename, so the period is resolved rather
# than written here. See scripts/lib/newest_release.py: a pinned filename keeps working
# while reading an older release than the one published.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
from newest_release import newest_release, period_phrase  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
XLSX = newest_release(ROOT / "data/raw/ho_visas_extra", "eu-settlement-scheme-datasets")
OUT = ROOT / "src/data/live/eu-settlement-scheme.json"


def _int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (TypeError, ValueError):
        return 0


def col_map(headers):
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ---- D02: Outcomes by year ----
    ws = wb["Data_EUSS_D02"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[2]]
    cm = col_map(headers)
    print(f"D02 headers: {headers[:10]}")

    annual_by_outcome = defaultdict(lambda: defaultdict(int))
    all_time_by_outcome = defaultdict(int)
    nationality_recent = defaultdict(int)

    c_year = cm.get("Outcome year")
    c_outcome = cm.get("Latest case outcome")
    c_count = cm.get("Count of applications concluded")
    c_nat = cm.get("Nationality")

    for r in rows[3:]:
        try:
            year = int(float(str(r[c_year])))
        except (TypeError, ValueError):
            continue
        n = _int(r[c_count])
        if n == 0:
            continue
        outcome = str(r[c_outcome]).strip() if c_outcome is not None and r[c_outcome] else ""
        annual_by_outcome[year][outcome] += n
        all_time_by_outcome[outcome] += n
        if year >= 2024 and c_nat is not None and r[c_nat]:
            nationality_recent[str(r[c_nat]).strip()] += n

    # ---- D03: Outcomes by application type ----
    ws = wb["Data_EUSS_D03"]
    rows3 = list(ws.iter_rows(values_only=True))
    headers3 = [str(c).strip() if c else "" for c in rows3[2]]
    cm3 = col_map(headers3)
    print(f"D03 headers: {headers3[:10]}")

    by_app_type = defaultdict(lambda: defaultdict(int))  # app_type -> outcome -> count
    c_year3 = cm3.get("Outcome year")
    c_outcome3 = cm3.get("Latest case outcome")
    c_type3 = cm3.get("Application type")
    c_count3 = cm3.get("Count of applications concluded")

    for r in rows3[3:]:
        try:
            year = int(float(str(r[c_year3])))
        except (TypeError, ValueError):
            continue
        n = _int(r[c_count3])
        if n == 0:
            continue
        outcome = str(r[c_outcome3]).strip() if c_outcome3 is not None and r[c_outcome3] else ""
        app_type = str(r[c_type3]).strip() if c_type3 is not None and r[c_type3] else ""
        if app_type:
            by_app_type[app_type][outcome] += n

    # Annual summary
    annual_summary = []
    for y in sorted(annual_by_outcome.keys()):
        d = annual_by_outcome[y]
        annual_summary.append({
            "year": y,
            "total": sum(d.values()),
            "by_outcome": dict(d),
            "settled_grants": d.get("Settled", 0),
            "pre_settled_grants": d.get("Pre-Settled", 0),
            "refused": d.get("Refused", 0),
            "invalid": d.get("Invalid", 0),
            "withdrawn_or_void": d.get("Withdrawn or Void", 0),
        })

    top_nationalities_recent = sorted(nationality_recent.items(), key=lambda x: -x[1])[:25]

    out = {
        "source": (
            "Home Office Immigration Statistics: EU Settlement Scheme detailed "
            f"datasets, {period_phrase(XLSX)} release. Sheets "
            "Data_EUSS_D02 (outcomes by year) and Data_EUSS_D03 (outcomes by "
            "application type)."
        ),
        "lastUpdated": "2026-05-27",
        "release_date": "2026-05-21",
        "caveat": (
            "The EU Settlement Scheme grants confirmed status to EU and EEA "
            "nationals already living in the UK. It does NOT count as "
            "'immigration' in ONS LTIM or Home Office visa-grant headlines. "
            "It still affects population denominators. The scheme officially "
            "closed for new applications on 30 June 2021 but late applications "
            "with valid reasons are still being processed, and existing pre-"
            "settled status holders apply for settled status after 5 years."
        ),
        "headline": {
            "total_concluded_all_time": sum(all_time_by_outcome.values()),
            "all_time_settled": all_time_by_outcome.get("Settled", 0),
            "all_time_pre_settled": all_time_by_outcome.get("Pre-Settled", 0),
            "all_time_refused": all_time_by_outcome.get("Refused", 0),
            "all_time_invalid": all_time_by_outcome.get("Invalid", 0),
            "all_time_withdrawn_or_void": all_time_by_outcome.get("Withdrawn or Void", 0),
            "latest_year": annual_summary[-1]["year"] if annual_summary else None,
            "latest_year_total": annual_summary[-1]["total"] if annual_summary else 0,
            "latest_year_settled": annual_summary[-1]["settled_grants"] if annual_summary else 0,
        },
        "all_time_by_outcome": dict(all_time_by_outcome),
        "annual": annual_summary,
        "by_application_type": {k: dict(v) for k, v in by_app_type.items()},
        "top_nationalities_recent": [
            {"nationality": n, "concluded_2024_2025_combined": c}
            for n, c in top_nationalities_recent
        ],
    }
    OUT.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {OUT}")
    print(f"All-time concluded: {sum(all_time_by_outcome.values()):,}")
    print(f"All-time settled: {all_time_by_outcome.get('Grant Settled', 0):,}")
    print(f"All-time pre-settled: {all_time_by_outcome.get('Grant Pre-Settled', 0):,}")
    print(f"\nAnnual settled grants:")
    for y in annual_summary[-7:]:
        print(f"  {y['year']}: {y['settled_grants']:,} settled / {y['pre_settled_grants']:,} pre-settled / {y['total']:,} total")
    print(f"\nTop 5 nationalities recent (2024-2025):")
    for n, c in top_nationalities_recent[:5]:
        print(f"  {n:25s} {c:>10,}")


if __name__ == "__main__":
    main()
