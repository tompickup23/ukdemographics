#!/usr/bin/env python3
"""
Parse the ONS Long-Term International Migration (LTIM) publication
spreadsheet and emit a clean JSON file holding the time series, the
revisions history, the uncertainty intervals, and the LTIM vs Home
Office visas vs DWP NINo triangulation.

Input:
  data/raw/ons_ltim/ltim-ye-dec-2025-may2026publication.xlsx

Output:
  src/data/live/ons-ltim.json

The headline numbers come from Table 1 (overall trend by nationality)
and Table 2 (uncertainty intervals on the latest period). Table 4a is
the non-EU+ breakdown by reason for migration. Table 5 is the revisions
history showing how each period's estimate has changed across
successive publications - this is the most useful single thing in the
release for telling readers how much the headline moves between
provisional, revised, and completed estimates. Tables 6a and 6b are
the ONS / Home Office / DWP triangulation.
"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "data/raw/ons_ltim/ltim-ye-dec-2025-may2026publication.xlsx"
OUT = ROOT / "src/data/live/ons-ltim.json"


def _int(v):
    """Parse an integer from a cell value; return None for blanks."""
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (TypeError, ValueError):
        return None


def _period(v):
    """Canonical 'YE Mmm YY' string."""
    if v is None:
        return None
    return str(v).strip()


def read_table_1(wb):
    """Table 1: Long-term international migration by flow and nationality."""
    ws = wb["1"]
    by_nationality = {"all": [], "british": [], "eu_plus": [], "non_eu_plus": []}
    # Headers at row 6, data from row 7
    for row in ws.iter_rows(min_row=7, values_only=True):
        flow = row[0]
        period = _period(row[1])
        if not flow or not period:
            continue
        rec = {
            "all": _int(row[2]),
            "british": _int(row[3]),
            "eu_plus": _int(row[4]),
            "non_eu_plus": _int(row[5]),
        }
        if flow.lower() == "immigration":
            for k in by_nationality:
                by_nationality[k].append({"period": period, "flow": "immigration", "value": rec[k]})
        elif flow.lower() == "emigration":
            for k in by_nationality:
                by_nationality[k].append({"period": period, "flow": "emigration", "value": rec[k]})
        elif "net" in flow.lower():
            for k in by_nationality:
                by_nationality[k].append({"period": period, "flow": "net_migration", "value": rec[k]})
    return by_nationality


def read_table_2(wb):
    """Table 2: Uncertainty intervals for the latest periods."""
    ws = wb["2"]
    out = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        flow = row[0]
        period = _period(row[1])
        est, lo, hi = _int(row[2]), _int(row[3]), _int(row[4])
        if not flow or not period:
            continue
        out.append({
            "flow": str(flow).lower().replace(" ", "_"),
            "period": period,
            "estimate": est,
            "lower_95": lo,
            "upper_95": hi,
        })
    return out


def read_table_4a(wb):
    """Table 4a: Non-EU+ by reason for migration (work / study / family / other / asylum)."""
    ws = wb["4a"]
    out = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        flow = row[0]
        period = _period(row[1])
        if not flow or not period:
            continue
        out.append({
            "flow": str(flow).lower().replace(" ", "_"),
            "period": period,
            "all_reasons": _int(row[2]),
            "work_all": _int(row[3]),
            "work_main": _int(row[4]),
            "work_dependant": _int(row[5]),
            "study_all": _int(row[6]),
            "study_main": _int(row[7]),
            "study_dependant": _int(row[8]),
            "family": _int(row[9]),
            "other": _int(row[10]),
            "asylum": _int(row[11]),
        })
    return out


def read_table_5(wb):
    """Table 5: Revisions to official LTIM estimates across publications."""
    ws = wb["5"]
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        period = _period(row[0])
        publication = _period(row[1])
        history = row[2]
        if not period or not publication:
            continue
        out.append({
            "period": period,
            "publication": publication,
            "estimate_stage": str(history).split("\n")[0].strip() if history else None,
            "immigration_all": _int(row[3]),
            "immigration_british": _int(row[4]),
            "immigration_eu_plus": _int(row[5]),
            "immigration_non_eu_plus": _int(row[6]),
            "emigration_all": _int(row[7]),
            "emigration_british": _int(row[8]),
            "emigration_eu_plus": _int(row[9]),
            "emigration_non_eu_plus": _int(row[10]),
            "net_migration_all": _int(row[11]),
        })
    return out


def read_table_6a(wb):
    """Table 6a: EU+ triangulation - LTIM vs Home Office visas vs DWP NINo."""
    ws = wb["6a"]
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        period = _period(row[0])
        if not period:
            continue
        out.append({
            "period": period,
            "eu_plus_visa_holder_ltim": _int(row[1]),
            "eu_plus_visas_granted_ho": _int(row[2]),
            "eu_plus_nino_allocations": _int(row[3]),
            "all_eu_plus_ltim": _int(row[4]),
        })
    return out


def read_table_6b(wb):
    """Table 6b: Non-EU+ triangulation - LTIM vs Home Office visas vs DWP NINo."""
    ws = wb["6b"]
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        period = _period(row[0])
        if not period:
            continue
        out.append({
            "period": period,
            "non_eu_plus_visa_holder_ltim": _int(row[1]),
            "non_eu_plus_visas_granted_ho": _int(row[2]),
            "non_eu_plus_nino_allocations": _int(row[3]),
        })
    return out


def read_table_8(wb):
    """Table 8: EU+ immigration broken down (Settled Status / Visa Holder / Irish)."""
    ws = wb["8"]
    out = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        flow = row[0]
        period = _period(row[1])
        if not flow or not period:
            continue
        out.append({
            "flow": str(flow).lower().replace(" ", "_"),
            "period": period,
            "all_eu_plus": _int(row[2]),
            "eu_settled_status": _int(row[3]),
            "eu_plus_visa_holder": _int(row[4]),
            "irish_nationals": _int(row[5]),
        })
    return out


def compute_headline(by_nationality_table_1, uncertainty_table_2):
    """Extract YE Dec 25 headline + YE Mar 23 peak + revisions of YE Dec 24.

    Periods carry suffixes: 'P' (provisional), 'R' (revised). Latest YE Dec 25
    is 'YE Dec 25 P', YE Mar 23 peak is the bare 'YE Mar 23' (completed),
    YE Dec 24 is 'YE Dec 24 R' (revised). Match against the period stem
    (without trailing P/R) so the function is robust to suffix drift.
    """
    def _stem(period):
        if not period:
            return period
        return period.replace(" P R", "").replace(" P", "").replace(" R", "").strip()

    def latest(rows, period, flow):
        target = _stem(period)
        for r in rows:
            if _stem(r["period"]) == target and r["flow"] == flow:
                return r["value"]
        return None

    def latest_unc(period, flow):
        target = _stem(period)
        for r in uncertainty_table_2:
            if _stem(r["period"]) == target and r["flow"] == flow:
                return r
        return None

    ye_dec_25 = "YE Dec 25"
    ye_mar_23 = "YE Mar 23"
    ye_dec_24 = "YE Dec 24"

    return {
        "ye_dec_25": {
            "label": ye_dec_25,
            "immigration_all": latest(by_nationality_table_1["all"], ye_dec_25, "immigration"),
            "emigration_all": latest(by_nationality_table_1["all"], ye_dec_25, "emigration"),
            "net_migration_all": latest(by_nationality_table_1["all"], ye_dec_25, "net_migration"),
            "net_migration_non_eu_plus": latest(
                by_nationality_table_1["non_eu_plus"], ye_dec_25, "net_migration"
            ),
            "immigration_uncertainty_95": latest_unc(ye_dec_25, "immigration"),
            "emigration_uncertainty_95": latest_unc(ye_dec_25, "emigration"),
            "net_migration_uncertainty_95": latest_unc(ye_dec_25, "net_migration"),
        },
        "ye_dec_24_revised": {
            "label": ye_dec_24,
            "immigration_all": latest(by_nationality_table_1["all"], ye_dec_24, "immigration"),
            "emigration_all": latest(by_nationality_table_1["all"], ye_dec_24, "emigration"),
            "net_migration_all": latest(by_nationality_table_1["all"], ye_dec_24, "net_migration"),
        },
        "ye_mar_23_peak": {
            "label": ye_mar_23,
            "immigration_all": latest(by_nationality_table_1["all"], ye_mar_23, "immigration"),
            "emigration_all": latest(by_nationality_table_1["all"], ye_mar_23, "emigration"),
            "net_migration_all": latest(by_nationality_table_1["all"], ye_mar_23, "net_migration"),
        },
    }


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    by_nationality = read_table_1(wb)
    uncertainty = read_table_2(wb)
    non_eu_by_reason = read_table_4a(wb)
    revisions = read_table_5(wb)
    triangulation_eu = read_table_6a(wb)
    triangulation_non_eu = read_table_6b(wb)
    eu_breakdown = read_table_8(wb)
    headline = compute_headline(by_nationality, uncertainty)

    out = {
        "source": (
            "ONS Long-term international migration, provisional: year ending "
            "December 2025 (released 21 May 2026). Tables 1, 2, 4a, 5, 6a, 6b, 8."
        ),
        "lastUpdated": "2026-05-27",
        "release_date": "2026-05-21",
        "period_covered": "YE Jun 2012 through YE Dec 2025",
        "methodology": (
            "ONS replaced the International Passenger Survey (IPS) with the "
            "Home Office Borders and Immigration Data (HOBID) for visa-holder "
            "nationals from YE June 2021 onwards, and the Registration and "
            "Population Interaction Database (RAPID) for British nationals "
            "from November 2025. Estimates before YE June 2021 use the older "
            "IPS-based method and are NOT directly comparable to current "
            "figures. British emigration estimates for YE December 2024 moved "
            "from 77,000 (old IPS estimate) to 257,000 (new RAPID estimate), "
            "a 180,000 upward revision driven by the methodology change. "
            "Migration Observatory: the 2024 revision 'results from a change "
            "of methodology, not a change in the underlying trend.'"
        ),
        "caveats": [
            "Uncertainty intervals around the YE Dec 2025 net migration headline "
            "are 145,000 to 197,000 (95 percent). ONS notes the interval does "
            "not include uncertainty from the emigration re-arrival adjustment, "
            "so the real spread is wider.",
            "Visa overstayers who do not claim asylum are assumed to have "
            "emigrated in the current method. Irregular migrants who do not "
            "claim asylum are not counted at all.",
            "Provisional estimates are routinely revised; YE Dec 2024 was "
            "revised down by 100,000 (-23 percent) versus its initial estimate.",
            "The YE Mar 2023 peak figure (944,000) was itself revised upward "
            "from previously published 906,000. Even historical reference "
            "points are not stable across publications.",
        ],
        "headline": headline,
        "time_series_by_nationality": by_nationality,
        "uncertainty_intervals": uncertainty,
        "non_eu_plus_by_reason": non_eu_by_reason,
        "revisions": revisions,
        "triangulation_eu_plus": triangulation_eu,
        "triangulation_non_eu_plus": triangulation_non_eu,
        "eu_plus_breakdown": eu_breakdown,
    }
    OUT.write_text(json.dumps(out, indent=2))
    print(f"Wrote {OUT}")

    def _fmt(v):
        return f"{v:,}" if v is not None else "None"

    print(f"\nHeadline (YE Dec 25):")
    h = headline["ye_dec_25"]
    print(f"  Net migration:  {_fmt(h['net_migration_all'])}")
    print(f"  Immigration:    {_fmt(h['immigration_all'])}")
    print(f"  Emigration:     {_fmt(h['emigration_all'])}")
    print(f"  Non-EU+ net:    {_fmt(h['net_migration_non_eu_plus'])}")
    unc = h.get("net_migration_uncertainty_95")
    if unc:
        print(f"  95 percent CI:  {_fmt(unc['lower_95'])} to {_fmt(unc['upper_95'])}")

    print(f"\nPeak (YE Mar 23):")
    p = headline["ye_mar_23_peak"]
    print(f"  Net migration:  {_fmt(p['net_migration_all'])}")
    print(f"  Immigration:    {_fmt(p['immigration_all'])}")

    print(f"\nLength of series:")
    for k, rows in by_nationality.items():
        n_periods = len({r["period"] for r in rows})
        print(f"  {k:15s}: {len(rows)} rows, {n_periods} periods")

    print(f"\nRevisions covered ({len(revisions)} entries):")
    for r in revisions[:6]:
        print(f"  {r['period']:10s} {r['publication']:8s}  net={r['net_migration_all']:>10,}")


if __name__ == "__main__":
    main()
