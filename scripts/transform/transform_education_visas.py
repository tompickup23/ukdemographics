#!/usr/bin/env python3
"""
Aggregate Home Office education visa grants by course level for UKD.

Student visas are now 47 percent of non-EU+ immigration in LTIM (the
dominant single category). The visa data lets us see the course level
mix: Bachelors vs Masters vs PhD vs Below Bachelors. The composition
matters for downstream story-telling about which courses British
universities are now using to fill their international student pipeline.

Input:
  data/raw/ho_visas_extra/education-visas-datasets-mar-2026.xlsx

Output:
  src/data/live/education-visas.json

Sheets:
  Data_Edu_D01: Applications (Year, Quarter, Nationality, Region,
                Visa type subgroup, Course level, Applications)
  Data_Edu_D02: Grants (same schema, count column 'Grants')

Source: Home Office Immigration Statistics, year ending March 2026
release (21 May 2026).
"""
import json
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "data/raw/ho_visas_extra/education-visas-datasets-mar-2026.xlsx"
OUT = ROOT / "src/data/live/education-visas.json"


def _int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (TypeError, ValueError):
        return 0


def col_map(headers):
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def get_headers(ws, max_check=4):
    for hr in range(1, max_check + 1):
        for row in ws.iter_rows(min_row=hr, max_row=hr, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if "Year" in cells and "Course level" in cells:
                return hr, cells
    return None, None


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # Grants (Edu_D02)
    ws = wb["Data_Edu_D02"]
    hdr_row, headers = get_headers(ws)
    cm = col_map(headers)
    print(f"Headers (row {hdr_row}): {headers[:8]}")

    annual_total = defaultdict(int)
    annual_by_level = defaultdict(lambda: defaultdict(int))
    annual_by_subgroup = defaultdict(lambda: defaultdict(int))
    nationality_2025 = defaultdict(int)
    level_by_nationality_2025 = defaultdict(lambda: defaultdict(int))

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        try:
            year = int(float(str(row[cm["Year"]])))
        except (TypeError, ValueError, IndexError):
            continue
        n = _int(row[cm["Grants"]])
        if n == 0:
            continue
        level = str(row[cm["Course level"]]).strip() if "Course level" in cm and row[cm["Course level"]] else ""
        subgroup = str(row[cm["Visa type subgroup"]]).strip() if "Visa type subgroup" in cm and row[cm["Visa type subgroup"]] else ""
        nat = str(row[cm["Nationality"]]).strip() if "Nationality" in cm and row[cm["Nationality"]] else ""

        annual_total[year] += n
        if level:
            annual_by_level[year][level] += n
        if subgroup:
            annual_by_subgroup[year][subgroup] += n
        if year == 2025:
            if nat:
                nationality_2025[nat] += n
            if nat and level:
                level_by_nationality_2025[nat][level] += n

    # Applications (Edu_D01)
    ws = wb["Data_Edu_D01"]
    hdr_row, headers = get_headers(ws)
    cm = col_map(headers)
    apps_annual = defaultdict(int)
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        try:
            year = int(float(str(row[cm["Year"]])))
        except (TypeError, ValueError, IndexError):
            continue
        apps_annual[year] += _int(row[cm["Applications"]])

    annual = []
    for y in sorted(annual_total.keys()):
        annual.append({
            "year": y,
            "grants": annual_total[y],
            "applications": apps_annual.get(y, 0),
            "grant_rate_pct": (
                round(annual_total[y] / apps_annual[y] * 100, 1)
                if apps_annual.get(y, 0) > 0 else None
            ),
            "by_level": dict(annual_by_level.get(y, {})),
            "by_subgroup": dict(annual_by_subgroup.get(y, {})),
        })

    top_nationalities_2025 = sorted(nationality_2025.items(), key=lambda x: -x[1])[:20]

    out = {
        "source": (
            "Home Office Immigration Statistics: Education visa applications "
            "and grants by course level, year ending March 2026 release "
            "(21 May 2026). Sheets Data_Edu_D01 (applications) and "
            "Data_Edu_D02 (grants of entry clearance for study)."
        ),
        "lastUpdated": "2026-05-27",
        "release_date": "2026-05-21",
        "caveat": (
            "Education visa grants include main applicants AND dependants of "
            "the main applicant. The January 2024 ban on most postgraduate "
            "study dependants is visible in the 2024 and 2025 dependant cohort "
            "fall. Course level reflects the level at the point of visa grant; "
            "actual enrolment can change after arrival."
        ),
        "headline": {
            "latest_year_grants": annual_total[max(annual_total.keys())] if annual_total else 0,
            "latest_year": max(annual_total.keys()) if annual_total else None,
        },
        "annual": annual,
        "top_nationalities_2025": [
            {"nationality": n, "grants": c} for n, c in top_nationalities_2025
        ],
        "level_by_top_nationalities_2025": {
            n: dict(level_by_nationality_2025[n])
            for n, _ in top_nationalities_2025[:10]
        },
    }
    OUT.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {OUT}")
    print(f"Annual grants (last 5 years):")
    for y in sorted(annual_total.keys())[-5:]:
        levels_str = " / ".join(f"{k}: {v:,}" for k, v in sorted(annual_by_level.get(y, {}).items(), key=lambda x: -x[1])[:4])
        print(f"  {y}: {annual_total[y]:,} grants. {levels_str}")
    print(f"\nTop 10 nationalities 2025:")
    for n, c in top_nationalities_2025[:10]:
        print(f"  {n:25s} {c:>8,}")


if __name__ == "__main__":
    main()
