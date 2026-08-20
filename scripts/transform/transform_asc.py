#!/usr/bin/env python3
"""
Transform NHS Digital ASCFR & SALT 2023/24 XLSX → asc-dashboard.json.

Adult Social Care responsibility sits with upper-tier authorities only
(counties, unitaries, London boroughs, mets). Districts (E07xxxxxx that aren't
unitarised) have no ASC budget and are correctly absent.

Tables used:
  T2  — summary (column 7+ are activity counts; col13 is number of clients
        receiving long-term support 18-64)
  T14 — gross current expenditure, col 13 = Total GCE in £k
  T34 — clients accessing long-term support during year — residential 18-64

DToC was discontinued post-COVID and is not in this file. ASCOF quality-of-life
sits in a separate publication and is also out of scope here. Both fields are
written as null with a caveat in the dashboard metadata.

Outputs: src/data/live/asc-dashboard.json
"""
from pathlib import Path
import json
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "data/raw/supplementary/nhs-asc-ascfr-salt-2023-24.xlsx"
OUT = ROOT / "src/data/live/asc-dashboard.json"
EP = ROOT / "src/data/live/ethnic-projections.json"


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def la_rows(ws, gss_col=2, name_col=4):
    """Return [(row, gss_code, la_name)] where gss starts with E0/E1."""
    out = []
    for r in range(8, ws.max_row + 1):
        gss = ws.cell(r, gss_col).value
        if isinstance(gss, str) and (gss.startswith("E06") or gss.startswith("E08") or gss.startswith("E09") or gss.startswith("E10")):
            out.append((r, gss, ws.cell(r, name_col).value or gss))
    return out


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    t14 = wb["T14"]  # spending
    t34 = wb["T34"]  # residential clients 18-64 + 65+

    ep = json.loads(EP.read_text())
    pop_by_la = {code: (a.get("current", {}).get("total_population") or 0) for code, a in ep["areas"].items()}
    name_by_la = {code: a.get("areaName", code) for code, a in ep["areas"].items()}

    # Spend per capita: T14 col 13 is Total GCE in £k.
    spend = {}
    for r, gss, name in la_rows(t14):
        gce_k = num(t14.cell(r, 13).value)
        pop = pop_by_la.get(gss) or 0
        if gce_k is not None and pop > 0:
            spend[gss] = (gce_k * 1000) / pop  # £k → £, then per head

    # Residential placement count for 65+: T34 has 18-64 then 65+ blocks.
    # Per the header inspection, columns shift: looking at headers programmatically
    # is safer than hard-coding. Find the 65+ Residential column by header text.
    h7 = [t34.cell(7, c).value for c in range(1, t34.max_column + 1)]
    h8 = [t34.cell(8, c).value for c in range(1, t34.max_column + 1)]
    # The "65 and over" header sits over a block. The 65+ Residential column
    # is the second "Residential" cell in the row-8 sub-headers.
    residential_cols = [c + 1 for c, v in enumerate(h8) if isinstance(v, str) and v.strip().lower().startswith("residential")]
    residential_65_col = residential_cols[1] if len(residential_cols) >= 2 else None

    residential_count = {}
    if residential_65_col:
        for r, gss, name in la_rows(t34):
            v = num(t34.cell(r, residential_65_col).value)
            if v is not None:
                residential_count[gss] = v

    # 65+ population denominator: rough — use 18.5% of total population
    # (England average from ONS 2022 mid-year estimates). Not exact but
    # gives a comparable rate across LAs given we lack age-band per LA in
    # ethnic-projections.json.
    POP_65_SHARE = 0.185

    areas = {}
    for r, gss, _ in la_rows(t14):
        spend_per_cap = spend.get(gss)
        rc = residential_count.get(gss)
        pop = pop_by_la.get(gss) or 0
        pop_65 = pop * POP_65_SHARE if pop else 0
        rate_65 = (rc / pop_65 * 10000) if (rc is not None and pop_65) else None
        areas[gss] = {
            "areaName": name_by_la.get(gss) or gss,
            "grossSpendPerCapita": round(spend_per_cap, 0) if spend_per_cap is not None else None,
            "residentialRatePer10k65": round(rate_65, 0) if rate_65 is not None else None,
            "qualityOfLifeScore": None,
            "dtocDaysAnnual": None,
            "period": "2023-24",
        }

    out = {
        "source": "NHS Digital ASCFR & SALT data tables 2023-24 (CASSR-level). Quality-of-life and DToC fields omitted (DToC discontinued post-COVID; ASCOF measures live in a separate publication).",
        "methodology": "Gross spend per capita = Total Gross Current Expenditure (T14 col 13, £k) × 1,000 ÷ LA total population (Census 2021). Residential placement rate per 10k 65+ = T34 65+ residential clients ÷ (total population × 0.185) × 10,000 — uses England-average 65+ share as denominator since per-LA age detail is not in this feed.",
        "lastUpdated": "2026-04-28",
        "caveat": "ASC sits with upper-tier authorities only (counties, unitaries, London boroughs, mets); ~153 LAs in coverage and districts are not present. Spending is shaped by demographic composition, deprivation, and informal-care availability and direct cross-area comparison must control for those.",
        "areas": dict(sorted(areas.items())),
    }
    OUT.write_text(json.dumps(out, indent=2))
    n = sum(1 for v in areas.values() if v["grossSpendPerCapita"] is not None)
    print(f"asc-dashboard.json: {len(areas)} CASSR areas ({n} with spend per capita)")


if __name__ == "__main__":
    main()
