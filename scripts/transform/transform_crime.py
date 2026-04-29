#!/usr/bin/env python3
"""
Transform ONS recorded crime XLSX (year ending Mar 2024) → crime-dashboard.json.
Source rows are CSP (Community Safety Partnership) level. Most CSPs map 1:1 to
LAs (~278 LAs); 14 LAs sit inside merged CSPs (Devon districts, Somerset
districts) and get the parent CSP's rates. ~17 LAs have no current CSP entry
(post-LGR reorganisations) — these stay omitted.

Outputs: src/data/live/crime-dashboard.json
"""
from pathlib import Path
import json
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "data/raw/supplementary/ons-recorded-crime-csp.xlsx"
OUT = ROOT / "src/data/live/crime-dashboard.json"
EP = ROOT / "src/data/live/ethnic-projections.json"


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    rates = wb["Table C5"]
    yoy = wb["Table C4"]
    asb = wb["Table C6"]
    crosswalk = wb["Table C1"]

    # Build CSP → [LA codes] from C1, picking up merged CSPs (rows where the
    # CSP cell is blank because openpyxl carries forward the previous value).
    csp_to_las = {}
    last_csp = None
    for r in range(5, crosswalk.max_row + 1):
        csp = crosswalk.cell(r, 2).value
        la_code = crosswalk.cell(r, 3).value
        if csp:
            last_csp = csp
        if la_code and isinstance(la_code, str) and la_code.startswith("E"):
            csp_to_las.setdefault(last_csp or "_unknown", []).append(la_code)

    merged_csps = {k: v for k, v in csp_to_las.items() if len(v) > 1}

    # Build LA → CSP-row index from C5 (column layout in head row 8).
    # col5=LA code, col6=LA name, col7=population, col9=total rate, col10=violent,
    # col17=robbery, col18=theft (umbrella), col19=burglary, col23=vehicle,
    # col26=shoplifting, col29=drug, col30=weapons, col31=public-order
    def yoy_row_for(la_code):
        for r in range(9, yoy.max_row + 1):
            if yoy.cell(r, 5).value == la_code:
                return r
        return None

    def num(v):
        return float(v) if isinstance(v, (int, float)) else None

    direct_rows = {}
    for r in range(9, rates.max_row + 1):
        la_code = rates.cell(r, 5).value
        if not (la_code and isinstance(la_code, str) and la_code.startswith("E")):
            continue
        la_name = (rates.cell(r, 6).value or "").strip()
        pop = num(rates.cell(r, 7).value) or 0
        total_rate = num(rates.cell(r, 9).value)
        violent_rate = num(rates.cell(r, 10).value)
        theft_rate = num(rates.cell(r, 18).value)
        drug_rate = num(rates.cell(r, 29).value)
        # ASB rate: derive from C6 count + C5 population
        asb_count = None
        for ar in range(9, asb.max_row + 1):
            if asb.cell(ar, 5).value == la_code:
                asb_count = num(asb.cell(ar, 7).value)
                break
        asb_rate = (asb_count / pop * 1000) if (asb_count is not None and pop) else None

        # Year-on-year on total recorded crime (C4 col 7)
        yr = yoy_row_for(la_code)
        yoy_pct = num(yoy.cell(yr, 7).value) if yr else None

        breakdown = []
        for label, col in [
            ("Violence against the person", 10),
            ("Sexual offences", 16),
            ("Robbery", 17),
            ("Theft offences", 18),
            ("Burglary", 19),
            ("Vehicle offences", 23),
            ("Drug offences", 29),
            ("Public order offences", 31),
        ]:
            v = num(rates.cell(r, col).value)
            if v is not None:
                breakdown.append({"type": label, "rate": round(v, 2)})

        direct_rows[la_code] = {
            "areaName": la_name,
            "totalCrimeRate": round(total_rate, 2) if total_rate is not None else None,
            "violentCrimeRate": round(violent_rate, 2) if violent_rate is not None else None,
            "theftRate": round(theft_rate, 2) if theft_rate is not None else None,
            "asbRate": round(asb_rate, 2) if asb_rate is not None else None,
            "drugRate": round(drug_rate, 2) if drug_rate is not None else None,
            "hateCrimeCount": None,
            "yearOnYearChange": round(yoy_pct, 2) if yoy_pct is not None else None,
            "breakdown": breakdown,
            "period": "Year ending March 2024",
            "source": "direct",
        }

    # Backfill LAs sitting inside a merged CSP — replicate the parent CSP's
    # rates onto each constituent LA and flag source="merged_csp".
    ep = json.loads(EP.read_text())
    ep_codes = set(ep["areas"].keys())
    def is_aggregate_la(v):
        return (not v) or (isinstance(v, str) and not v.startswith("E"))

    for csp_name, las in merged_csps.items():
        # Find the CSP-level row in C5 (csp name col 4, LA cell is "Combined ..." or null)
        for r in range(9, rates.max_row + 1):
            if rates.cell(r, 4).value != csp_name or not is_aggregate_la(rates.cell(r, 5).value):
                continue
            pop = num(rates.cell(r, 7).value) or 0
            total_rate = num(rates.cell(r, 9).value)
            violent_rate = num(rates.cell(r, 10).value)
            theft_rate = num(rates.cell(r, 18).value)
            drug_rate = num(rates.cell(r, 29).value)
            asb_count = None
            for ar in range(9, asb.max_row + 1):
                if asb.cell(ar, 4).value == csp_name and is_aggregate_la(asb.cell(ar, 5).value):
                    asb_count = num(asb.cell(ar, 7).value)
                    break
            asb_rate = (asb_count / pop * 1000) if (asb_count is not None and pop) else None
            breakdown = []
            for label, col in [
                ("Violence against the person", 10),
                ("Sexual offences", 16),
                ("Robbery", 17),
                ("Theft offences", 18),
                ("Burglary", 19),
                ("Vehicle offences", 23),
                ("Drug offences", 29),
                ("Public order offences", 31),
            ]:
                v = num(rates.cell(r, col).value)
                if v is not None:
                    breakdown.append({"type": label, "rate": round(v, 2)})
            for la in las:
                if la in direct_rows or la not in ep_codes:
                    continue
                direct_rows[la] = {
                    "areaName": ep["areas"][la].get("areaName", la),
                    "totalCrimeRate": round(total_rate, 2) if total_rate is not None else None,
                    "violentCrimeRate": round(violent_rate, 2) if violent_rate is not None else None,
                    "theftRate": round(theft_rate, 2) if theft_rate is not None else None,
                    "asbRate": round(asb_rate, 2) if asb_rate is not None else None,
                    "drugRate": round(drug_rate, 2) if drug_rate is not None else None,
                    "hateCrimeCount": None,
                    "yearOnYearChange": None,
                    "breakdown": breakdown,
                    "period": "Year ending March 2024",
                    "source": f"shared_csp:{csp_name}",
                }
            break

    out = {
        "source": "ONS recorded crime by Community Safety Partnership area, year ending March 2024 (Home Office police recorded crime). LA-level rates are CSP rates inherited where multiple LAs share a CSP.",
        "methodology": "Total recorded crime rate per 1,000 population, year ending March 2024. Where a CSP groups multiple LAs (e.g. Devon districts) constituent LAs inherit the CSP rate. ASB rate derived from Table C6 count divided by Table C5 population. Year-on-year is total recorded crime % change vs YE March 2023.",
        "lastUpdated": "2026-04-28",
        "caveat": "Police recorded crime is shaped by recording practice, reporting rates, and policing priority. Cross-area comparison must take account of those factors. Hate crime and quality-of-life detail are not in this file.",
        "areas": dict(sorted(direct_rows.items())),
    }
    OUT.write_text(json.dumps(out, indent=2))
    direct = sum(1 for v in direct_rows.values() if v["source"] == "direct")
    inherited = sum(1 for v in direct_rows.values() if v["source"].startswith("shared_csp"))
    print(f"crime-dashboard.json: {len(direct_rows)} LAs ({direct} direct + {inherited} inherited from merged CSPs)")


if __name__ == "__main__":
    main()
