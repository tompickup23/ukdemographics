#!/usr/bin/env python3
"""
Aggregate Home Office work visa grants by SOC2020 occupation for UKD.

The single biggest story behind the YE Mar 2026 net migration fall is
the Health and Care Worker visa collapse - 108,000 main applicants in
Caring Personal Service roles fell to 1,400. This file surfaces that
story plus broader sector composition: which occupations dominate work
visa grants, how they have shifted, and which industries the workers
are deployed in.

Input:
  data/raw/ho_visas_extra/occupation-soc2020-visas-datasets-mar-2026.xlsx

Output:
  src/data/live/work-visas-by-occupation.json

Sheets:
  Data_Occ_D01: Applications (year, nationality, region, visa type,
    visa subgroup, industry, SOC major/sub-major/minor/unit, count)
  Data_Occ_D02: Grants (same schema as D01, count column 'Grants')

Source: Home Office, year ending March 2026 release (21 May 2026).
"""
import json
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "data/raw/ho_visas_extra/occupation-soc2020-visas-datasets-mar-2026.xlsx"
OUT = ROOT / "src/data/live/work-visas-by-occupation.json"


def _int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (TypeError, ValueError):
        return 0


def find_header_row(ws, expected_cols, max_check=6):
    """Find the row index containing the expected column names."""
    for hr in range(1, max_check + 1):
        for row in ws.iter_rows(min_row=hr, max_row=hr, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if all(any(ec.lower() in c.lower() for c in cells) for ec in expected_cols):
                return hr, cells
    return None, None


def read_data(ws, header_row, header_cells, count_col_name):
    """Yield dicts of {column: value} for each data row."""
    # Map header name to column index. Handle the count column matching.
    col_map = {}
    for i, h in enumerate(header_cells):
        if h:
            col_map[h] = i

    count_idx = None
    for k, v in col_map.items():
        if k.lower() == count_col_name.lower():
            count_idx = v
            break

    if count_idx is None:
        raise ValueError(f"Count column '{count_col_name}' not found in {list(col_map.keys())}")

    year_idx = col_map.get("Year")
    nat_idx = col_map.get("Nationality")
    region_idx = col_map.get("Region")
    visa_type_idx = col_map.get("Visa type")
    visa_sub_idx = col_map.get("Visa type subgroup")
    industry_idx = col_map.get("Industry")
    major_idx = col_map.get("Occ. major group")
    minor_idx = col_map.get("Occ. minor group")
    unit_idx = col_map.get("Occ. unit group")

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) <= count_idx:
            continue
        n = _int(row[count_idx])
        if n == 0:
            continue
        yield {
            "year": int(float(str(row[year_idx]))) if year_idx is not None and row[year_idx] else None,
            "nationality": str(row[nat_idx]).strip() if nat_idx is not None and row[nat_idx] else "",
            "region": str(row[region_idx]).strip() if region_idx is not None and row[region_idx] else "",
            "visa_type": str(row[visa_type_idx]).strip() if visa_type_idx is not None and row[visa_type_idx] else "",
            "visa_subgroup": str(row[visa_sub_idx]).strip() if visa_sub_idx is not None and row[visa_sub_idx] else "",
            "industry": str(row[industry_idx]).strip() if industry_idx is not None and row[industry_idx] else "",
            "soc_major": str(row[major_idx]).strip() if major_idx is not None and row[major_idx] else "",
            "soc_minor": str(row[minor_idx]).strip() if minor_idx is not None and row[minor_idx] else "",
            "soc_unit": str(row[unit_idx]).strip() if unit_idx is not None and row[unit_idx] else "",
            "count": n,
        }


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    ws_grants = wb["Data_Occ_D02"]
    hdr_row, hdr_cells = find_header_row(ws_grants, ["Year", "Nationality", "Grants"])
    print(f"Grants header row {hdr_row}: {hdr_cells[:12]}")

    annual_total = defaultdict(int)                                     # year -> total
    annual_by_subgroup = defaultdict(lambda: defaultdict(int))          # year -> subgroup -> total
    annual_by_industry = defaultdict(lambda: defaultdict(int))          # year -> industry -> total
    annual_by_major = defaultdict(lambda: defaultdict(int))             # year -> SOC major -> total
    annual_by_minor = defaultdict(lambda: defaultdict(int))             # year -> SOC minor -> total
    annual_by_unit = defaultdict(lambda: defaultdict(int))              # year -> SOC unit -> total
    by_nationality_2025 = defaultdict(int)                              # nat -> count
    by_subgroup_2025 = defaultdict(int)
    care_workers_by_year = defaultdict(int)                             # year -> care_worker count

    total_processed = 0
    for r in read_data(ws_grants, hdr_row, hdr_cells, "Grants"):
        if r["year"] is None:
            continue
        n = r["count"]
        y = r["year"]
        annual_total[y] += n
        if r["visa_subgroup"]:
            annual_by_subgroup[y][r["visa_subgroup"]] += n
        if r["industry"]:
            annual_by_industry[y][r["industry"]] += n
        if r["soc_major"]:
            annual_by_major[y][r["soc_major"]] += n
        if r["soc_minor"]:
            annual_by_minor[y][r["soc_minor"]] += n
        if r["soc_unit"]:
            annual_by_unit[y][r["soc_unit"]] += n
        # Care worker tracking - match SOC unit 6135 "Care workers and home carers"
        # and minor group 613 "Caring personal services"
        unit = r["soc_unit"].lower()
        minor = r["soc_minor"].lower()
        if unit.startswith("6135") or "caring personal" in minor or "care workers" in unit:
            care_workers_by_year[y] += n
        if y == 2025:
            if r["nationality"]:
                by_nationality_2025[r["nationality"]] += n
            if r["visa_subgroup"]:
                by_subgroup_2025[r["visa_subgroup"]] += n
        total_processed += n

    # Top SOC units across all years
    soc_unit_grand_total = defaultdict(int)
    for y_dict in annual_by_unit.values():
        for unit, c in y_dict.items():
            soc_unit_grand_total[unit] += c

    # Top SOC units by 2025 share
    top_units_2025 = sorted(annual_by_unit.get(2025, {}).items(), key=lambda x: -x[1])[:25]
    top_units_alltime = sorted(soc_unit_grand_total.items(), key=lambda x: -x[1])[:25]

    # Annual time series for top 10 SOC units
    top_10_units = [u for u, _ in top_units_alltime[:10]]
    unit_time_series = {}
    for unit in top_10_units:
        series = {}
        for y in sorted(annual_by_unit.keys()):
            series[y] = annual_by_unit[y].get(unit, 0)
        unit_time_series[unit] = series

    out = {
        "source": (
            "Home Office Immigration Statistics: Occupation by SOC2020 visas, "
            "year ending March 2026 release (21 May 2026). Sheet Data_Occ_D02 "
            "(grants of entry clearance work visas)."
        ),
        "lastUpdated": "2026-05-27",
        "release_date": "2026-05-21",
        "caveat": (
            "Work visa grants are issued at the entry-clearance stage; not "
            "every grant leads to long-term migration. The Health and Care "
            "Worker route's near-collapse (108,000 main applicants in Caring "
            "Personal Service roles fell to 1,400 between YE Dec 2023 and YE "
            "Mar 2026) is the single biggest driver of the work-visa fall. "
            "Skilled Worker visas overall fell 76 percent from their YE Dec "
            "2023 peak. Dependants are included unless filtered by visa type."
        ),
        "headline": {
            "total_grants_2025": annual_total.get(2025, 0),
            "total_grants_2024": annual_total.get(2024, 0),
            "total_grants_2023": annual_total.get(2023, 0),
            "yoy_change_2024_2025_pct": (
                round((annual_total.get(2025, 0) - annual_total.get(2024, 0))
                      / max(annual_total.get(2024, 1), 1) * 100, 1)
            ),
            "care_workers_peak_year": max(care_workers_by_year, key=care_workers_by_year.get) if care_workers_by_year else None,
            "care_workers_peak_count": max(care_workers_by_year.values()) if care_workers_by_year else 0,
            "care_workers_2025": care_workers_by_year.get(2025, 0),
        },
        "annual_total": dict(annual_total),
        "annual_by_subgroup": {y: dict(d) for y, d in annual_by_subgroup.items()},
        "annual_by_industry": {y: dict(d) for y, d in annual_by_industry.items()},
        "annual_by_soc_major": {y: dict(d) for y, d in annual_by_major.items()},
        "top_soc_units_2025": [{"soc_unit": u, "grants": c} for u, c in top_units_2025],
        "top_soc_units_alltime": [{"soc_unit": u, "grants": c} for u, c in top_units_alltime],
        "top_nationalities_2025": [
            {"nationality": n, "grants": c}
            for n, c in sorted(by_nationality_2025.items(), key=lambda x: -x[1])[:20]
        ],
        "top_subgroups_2025": [
            {"subgroup": s, "grants": c}
            for s, c in sorted(by_subgroup_2025.items(), key=lambda x: -x[1])
        ],
        "care_workers_by_year": dict(care_workers_by_year),
        "time_series_top_10_units": unit_time_series,
    }
    OUT.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {OUT}")
    print(f"Total processed (grants): {total_processed:,}")
    print(f"\nAnnual totals (last 5 years):")
    for y in sorted(annual_total.keys())[-5:]:
        print(f"  {y}: {annual_total[y]:,}")
    print(f"\nCare workers trajectory:")
    for y in sorted(care_workers_by_year.keys())[-6:]:
        print(f"  {y}: {care_workers_by_year[y]:,}")
    print(f"\nTop 10 SOC units (2025 grants):")
    for u in top_units_2025[:10]:
        print(f"  {u[0]:60s} {u[1]:>8,}")
    print(f"\nTop 10 nationalities (2025 work grants):")
    for n, c in sorted(by_nationality_2025.items(), key=lambda x: -x[1])[:10]:
        print(f"  {n:25s} {c:>8,}")


if __name__ == "__main__":
    main()
