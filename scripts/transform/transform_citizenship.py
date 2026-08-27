#!/usr/bin/env python3
"""
Aggregate Home Office British citizenship grants for UKD.

Naturalisation grants are the tail-end of the migration journey:
people who arrived years earlier, settled, and have now formally
become British citizens. The volume and nationality mix tells you who
the long-term stayers were a decade ago, which is useful counter-
weight to current-year visa-grant churn.

Input:
  data/raw/ho_visas_extra/citizenship-datasets-mar-2026.xlsx

Output:
  src/data/live/citizenship.json

Sheets:
  Data_Cit_D01: Applications (Year, Quarter, App type group, Region, Nationality, Applications)
  Data_Cit_D02: Grants (Year, Quarter, App type group, App type, Region, Nationality, Sex, Age, Grants)
  Data_Cit_D03: Ceremonies attended (Year, UK Region, Local authority, Ceremonies attended)

Source: Home Office Immigration Statistics, year ending March 2026
release (21 May 2026).
"""
import json
from collections import defaultdict
import sys
from pathlib import Path

import openpyxl

# Release files carry their period in the filename, so the period is resolved rather
# than written here. See scripts/lib/newest_release.py: a pinned filename keeps working
# while reading an older release than the one published.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
from newest_release import newest_release, period_phrase  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
XLSX = newest_release(ROOT / "data/raw/ho_visas_extra", "citizenship-datasets")
OUT = ROOT / "src/data/live/citizenship.json"


def _int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (TypeError, ValueError):
        return 0


def col_map(headers):
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None and str(h).strip()}


def get_headers(ws, max_check=4):
    for hr in range(1, max_check + 1):
        for row in ws.iter_rows(min_row=hr, max_row=hr, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if "Year" in cells:
                return hr, cells
    return None, None


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # ---- Grants (Cit_D02) ----
    ws = wb["Data_Cit_D02"]
    hdr_row, headers = get_headers(ws)
    cm = col_map(headers)
    print(f"Cit_D02 headers (row {hdr_row}): {headers[:10]}")

    annual_total = defaultdict(int)
    annual_by_type = defaultdict(lambda: defaultdict(int))   # year -> app type group -> count
    annual_by_nationality_2024 = defaultdict(int)
    annual_by_nationality_2025 = defaultdict(int)
    all_time_by_nationality = defaultdict(int)

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        try:
            year = int(float(str(row[cm["Year"]])))
        except (TypeError, ValueError, IndexError):
            continue
        n = _int(row[cm["Grants"]])
        if n == 0:
            continue
        type_group = str(row[cm["Application type group"]]).strip() if "Application type group" in cm else ""
        nat = str(row[cm["Nationality"]]).strip() if "Nationality" in cm else ""

        annual_total[year] += n
        if type_group:
            annual_by_type[year][type_group] += n
        if nat:
            all_time_by_nationality[nat] += n
            if year == 2024:
                annual_by_nationality_2024[nat] += n
            elif year == 2025:
                annual_by_nationality_2025[nat] += n

    # ---- Applications (Cit_D01) ----
    ws = wb["Data_Cit_D01"]
    hdr_row, headers = get_headers(ws)
    cm = col_map(headers)
    apps_annual_total = defaultdict(int)
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        try:
            year = int(float(str(row[cm["Year"]])))
        except (TypeError, ValueError, IndexError):
            continue
        n = _int(row[cm["Applications"]])
        apps_annual_total[year] += n

    # Top nationalities by 2025 + 2024 combined (more stable)
    combined_nat = defaultdict(int)
    for n, c in annual_by_nationality_2024.items():
        combined_nat[n] += c
    for n, c in annual_by_nationality_2025.items():
        combined_nat[n] += c
    top_nationalities = sorted(combined_nat.items(), key=lambda x: -x[1])[:20]

    annual_summary = []
    for y in sorted(annual_total.keys()):
        annual_summary.append({
            "year": y,
            "applications": apps_annual_total.get(y, 0),
            "grants": annual_total[y],
            "grant_rate_pct": (
                round(annual_total[y] / apps_annual_total[y] * 100, 1)
                if apps_annual_total.get(y, 0) > 0 else None
            ),
            "by_type": dict(annual_by_type.get(y, {})),
        })

    out = {
        "source": (
            "Home Office Immigration Statistics: Citizenship detailed datasets, "
            f"{period_phrase(XLSX)} release. Sheets Cit_D01 "
            "(applications) and Cit_D02 (grants)."
        ),
        "lastUpdated": "2026-05-27",
        "release_date": "2026-05-21",
        "caveat": (
            "Citizenship grants are a lagging indicator of migration: most "
            "naturalisations are granted to people who arrived in the UK "
            "more than five years earlier and have since settled. The volume "
            "reflects who arrived a decade ago, NOT current migration policy."
        ),
        "headline": {
            "latest_year_grants": annual_total[max(annual_total.keys())] if annual_total else 0,
            "latest_year": max(annual_total.keys()) if annual_total else None,
            "all_time_total_grants": sum(annual_total.values()),
            "all_time_total_applications": sum(apps_annual_total.values()),
        },
        "annual": annual_summary,
        "top_nationalities_recent_2y": [
            {"nationality": n, "grants_2024_2025_combined": c} for n, c in top_nationalities
        ],
        "all_time_top_nationalities": [
            {"nationality": n, "all_time_grants": c}
            for n, c in sorted(all_time_by_nationality.items(), key=lambda x: -x[1])[:30]
        ],
    }
    OUT.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {OUT}")
    print(f"Annual grants 2020-2025:")
    for y in sorted(annual_total.keys()):
        if y >= 2020:
            print(f"  {y}: {annual_total[y]:,} grants, {apps_annual_total.get(y, 0):,} applications")
    print(f"\nTop 10 nationalities by 2024-2025 combined grants:")
    for n, c in top_nationalities[:10]:
        print(f"  {n:25s} {c:>8,}")


if __name__ == "__main__":
    main()
